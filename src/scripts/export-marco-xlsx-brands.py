"""Export SKU/name -> brand maps from Marco.xlsx for backfill script."""
import json
import sys
from pathlib import Path

from openpyxl import load_workbook

XLSX_PATH = Path(r"c:\Users\ROG\Downloads\Telegram Desktop\Marco.xlsx")
OUT_PATH = Path(__file__).resolve().parent / "marco-xlsx-brand-map.json"


def normalize_sku(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def normalize_brand(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def infer_brand_from_name(name: str) -> str:
    name = name.strip()
    if not name:
        return ""
    # Title often starts with brand: "Samsung QA55...", "Welux W0 60"
    return name.split()[0].strip()


def main() -> None:
    wb = load_workbook(XLSX_PATH, read_only=True, data_only=True)
    ws = wb.active

    header = None
    sku_idx = brand_idx = name_idx = -1
    by_sku: dict[str, str] = {}
    by_name: dict[str, str] = {}

    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            header = row
            sku_idx = header.index("Артикул")
            brand_idx = header.index("Brand")
            name_idx = header.index("Name")
            continue

        if not row or not any(row):
            continue

        sku = normalize_sku(row[sku_idx])
        brand = normalize_brand(row[brand_idx])
        name = normalize_brand(row[name_idx])

        if not sku and not name:
            continue

        resolved = brand or infer_brand_from_name(name)
        if not resolved:
            continue

        if sku:
            by_sku[sku] = resolved
        if name:
            by_name[name.casefold()] = resolved

    payload = {
        "source": str(XLSX_PATH),
        "bySku": by_sku,
        "byName": by_name,
        "stats": {
            "skuMappings": len(by_sku),
            "nameMappings": len(by_name),
        },
    }

    OUT_PATH.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    sys.stdout.write(f"Wrote {OUT_PATH} ({payload['stats']})\n")


if __name__ == "__main__":
    main()
