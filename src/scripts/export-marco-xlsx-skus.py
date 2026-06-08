"""Export product title -> articul (SKU) mappings from Marco.xlsx."""
import json
import sys
from pathlib import Path

from openpyxl import load_workbook

DEFAULT_XLSX = Path(r"c:\Users\ROG\Downloads\Marco.xlsx")
OUT_PATH = Path(__file__).resolve().parent / "marco-xlsx-sku-map.json"
ARTICUL_HEADERS = ("Артикул", "Արտիկուլ", "Articul", "SKU")
NAME_HEADERS = ("Name", "Անուն", "name")


def normalize_sku(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def normalize_name(value) -> str:
    if value is None:
        return ""
    return " ".join(str(value).split()).strip()


def name_key(value: str) -> str:
    return normalize_name(value).casefold()


def find_index(header: tuple, candidates: tuple[str, ...]) -> int:
    for candidate in candidates:
        try:
            return header.index(candidate)
        except ValueError:
            continue
    return -1


def main() -> None:
    xlsx_path = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_XLSX
    if not xlsx_path.is_file():
        raise SystemExit(f"File not found: {xlsx_path}")

    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb.active

    entries: list[dict[str, str]] = []
    by_name: dict[str, list[str]] = {}
    by_articul: dict[str, str] = {}
    rows_total = 0
    rows_with_articul = 0

    header = None
    name_idx = articul_idx = -1

    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            header = row
            name_idx = find_index(header, NAME_HEADERS)
            articul_idx = find_index(header, ARTICUL_HEADERS)
            if name_idx < 0:
                raise SystemExit(f"Name column not found in: {header}")
            if articul_idx < 0:
                raise SystemExit(f"Articul column not found in: {header}")
            continue

        if not row or not any(row):
            continue

        name = normalize_name(row[name_idx])
        articul = normalize_sku(row[articul_idx])
        if not name:
            continue

        rows_total += 1
        if not articul:
            continue

        rows_with_articul += 1
        key = name_key(name)
        entries.append({"nameKey": key, "name": name, "articul": articul})
        by_name.setdefault(key, [])
        if articul not in by_name[key]:
            by_name[key].append(articul)
        by_articul[articul] = key

    payload = {
        "source": str(xlsx_path),
        "entries": entries,
        "byName": by_name,
        "byArticul": by_articul,
        "stats": {
            "rowsWithName": rows_total,
            "rowsWithArticul": rows_with_articul,
            "uniqueNames": len(by_name),
            "uniqueArticuls": len(by_articul),
        },
    }

    OUT_PATH.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Wrote {OUT_PATH} ({len(entries)} entries, {len(by_name)} unique names)")


if __name__ == "__main__":
    main()
