"""Build a deterministic product-attribute repair manifest from Marco.xlsx.

The manifest is intentionally independent from database state. Product SKUs are
identifiers (not quantities), so numeric cells using a zero-padding number format
retain their displayed leading zeroes.
"""

from __future__ import annotations

import argparse
import hashlib
import json
import re
from pathlib import Path
from typing import Any

import openpyxl


SCHEMA_VERSION = 1
COLOR_HEADERS = {"color", "գույն", "цвет"}
SKU_HEADERS = ("SKU", "Артикул", "Արտիկուլ", "Product code", "Product Code")
NAME_HEADERS = ("Name", "ƒ")

# These columns describe the product/import operation rather than a selectable
# product attribute. Every other non-empty Worksheet header is technical data.
NON_ATTRIBUTE_HEADERS = {
    "",
    "id",
    "name",
    "ƒ",
    "sku",
    "артикул",
    "արտիկուլ",
    "product code",
    "տիպ մեծածախ / մանրածախ",
    "տիպ մեծածախ /  մանրածախ",
    "тип մեծածախ / մանրածախ",
    "тип մեծածախ /  մանրածախ",
    "type",
    "draftstatus",
    "черновик",
    "short description",
    "description",
    "desctriptop",
    "price",
    "sale price",
    "category",
    "brand",
    "երաշխիք",
    "warranty",
    "images",
    "stock",
}

# `convert_marco_xlsx_for_import.py` historically numbered FilterN columns with
# this exact exclusion set. Some non-attribute source columns (notably Stock and
# Short description) were therefore counted. Keep a separate legacy counter so
# `marco_filter_N` compatibility does not drift while the repair manifest itself
# includes only real attributes.
LEGACY_FILTER_EXCLUDED_HEADERS = {
    "",
    "id",
    "name",
    "ƒ",
    "sku",
    "артикул",
    "արտիկուլ",
    "տիպ մեծածախ / մանրածախ",
    "тип մեծածախ / մանրածախ",
    "type",
    "draftstatus",
    "черновик",
    "description",
    "desctriptop",
    "price",
    "sale price",
    "category",
    "brand",
    "երաշխիք",
    "warranty",
    "color",
    "images",
}


def normalize_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return re.sub(r"\s+", " ", str(value).strip())


def normalize_header(value: object) -> str:
    return normalize_text(value)


def normalized_header_key(value: object) -> str:
    return normalize_header(value).casefold()


def identifier_from_cell(cell: Any) -> str:
    """Return an identifier using simple Excel zero-padding when present."""

    value = cell.value
    if value is None:
        return ""
    if isinstance(value, bool):
        return normalize_text(value)
    if isinstance(value, (int, float)) and float(value).is_integer():
        integer = int(value)
        number_format = str(getattr(cell, "number_format", "") or "")
        # Covers identifier formats used by the source (for example 00000).
        # Ignore decimal/scientific/date-like formats rather than guessing.
        sections = number_format.split(";", 1)[0]
        if re.fullmatch(r"0+", sections):
            return f"{integer:0{len(sections)}d}"
        return str(integer)
    return normalize_text(value)


def sha1_text(value: str, length: int = 10) -> str:
    return hashlib.sha1(value.encode("utf-8")).hexdigest()[:length]


def semantic_attribute_key(label: str, filter_index: int) -> str:
    """Mirror the current CSV importer's ASCII key derivation."""

    slug = re.sub(r"[^a-z0-9]+", "-", label.lower()).strip("-")[:80]
    return slug or f"spec-{filter_index}-{sha1_text(label)}"


def find_header_index(headers: list[str], candidates: tuple[str, ...]) -> int:
    for candidate in candidates:
        try:
            return headers.index(candidate)
        except ValueError:
            continue
    raise ValueError(f"Required Worksheet header not found; expected one of: {', '.join(candidates)}")


def build_attribute_definitions(headers: list[str]) -> list[dict[str, object]]:
    definitions: list[dict[str, object]] = []
    filter_index = 0
    legacy_filter_indices: dict[int, int] = {}
    legacy_filter_index = 0
    for column_index, header in enumerate(headers, start=1):
        if header and normalized_header_key(header) not in LEGACY_FILTER_EXCLUDED_HEADERS:
            legacy_filter_index += 1
            legacy_filter_indices[column_index] = legacy_filter_index

    for column_index, header in enumerate(headers, start=1):
        key = normalized_header_key(header)
        if not header or key in NON_ATTRIBUTE_HEADERS:
            continue
        if key in COLOR_HEADERS:
            definitions.append(
                {
                    "id": "color",
                    "kind": "color",
                    "label": header,
                    "sourceColumn": column_index,
                    "semanticKey": "color",
                    "compatibilityKeys": ["color"],
                }
            )
            continue

        filter_index += 1
        compatibility_index = legacy_filter_indices.get(column_index, filter_index)
        definitions.append(
            {
                "id": f"technical:{filter_index}",
                "kind": "technical",
                "label": header,
                "sourceColumn": column_index,
                "filterIndex": filter_index,
                "legacyFilterIndex": compatibility_index,
                "semanticKey": semantic_attribute_key(header, compatibility_index),
                "compatibilityKeys": [f"marco_filter_{compatibility_index}"],
            }
        )

    color_definitions = [item for item in definitions if item["kind"] == "color"]
    if len(color_definitions) > 1:
        columns = ", ".join(str(item["sourceColumn"]) for item in color_definitions)
        raise ValueError(f"Worksheet contains multiple Color columns: {columns}")
    return definitions


def build_manifest(xlsx_path: Path, sheet_name: str = "Worksheet") -> dict[str, object]:
    workbook = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    try:
        if sheet_name not in workbook.sheetnames:
            raise ValueError(
                f"Sheet '{sheet_name}' not found. Available: {', '.join(workbook.sheetnames)}"
            )
        worksheet = workbook[sheet_name]
        rows = worksheet.iter_rows(min_row=1)
        header_cells = next(rows, None)
        if header_cells is None:
            raise ValueError(f"Sheet '{sheet_name}' is empty")

        headers = [normalize_header(cell.value) for cell in header_cells]
        sku_index = find_header_index(headers, SKU_HEADERS)
        try:
            name_index = find_header_index(headers, NAME_HEADERS)
        except ValueError:
            name_index = -1
        definitions = build_attribute_definitions(headers)

        entries: list[dict[str, object]] = []
        skipped_missing_sku = 0
        worksheet_rows = 0
        seen_skus: dict[str, int] = {}

        for row_number, cells in enumerate(rows, start=2):
            if not any(normalize_text(cell.value) for cell in cells):
                continue
            worksheet_rows += 1
            sku = identifier_from_cell(cells[sku_index])
            if not sku:
                skipped_missing_sku += 1
                continue
            if sku in seen_skus:
                raise ValueError(
                    f"Duplicate SKU '{sku}' in Worksheet rows {seen_skus[sku]} and {row_number}"
                )
            seen_skus[sku] = row_number

            values: list[dict[str, object]] = []
            for definition in definitions:
                column_index = int(definition["sourceColumn"]) - 1
                if column_index >= len(cells):
                    continue
                value = normalize_text(cells[column_index].value)
                # A styled/formatted cell without a value is intentionally blank.
                if not value:
                    continue
                values.append(
                    {
                        "definitionId": definition["id"],
                        "value": value,
                        "sourceCell": cells[column_index].coordinate,
                    }
                )

            entries.append(
                {
                    "sku": sku,
                    "sourceRow": row_number,
                    "name": normalize_text(cells[name_index].value) if name_index >= 0 else "",
                    "values": values,
                }
            )

        entries.sort(key=lambda item: (str(item["sku"]), int(item["sourceRow"])))
        attribute_cells = sum(len(item["values"]) for item in entries)
        manifest: dict[str, object] = {
            "schemaVersion": SCHEMA_VERSION,
            "source": {
                "fileName": xlsx_path.name,
                "sha256": hashlib.sha256(xlsx_path.read_bytes()).hexdigest(),
                "sheet": sheet_name,
            },
            "attributeDefinitions": definitions,
            "entries": entries,
            "stats": {
                "worksheetRows": worksheet_rows,
                "entries": len(entries),
                "entriesWithValues": sum(1 for item in entries if item["values"]),
                "attributeCells": attribute_cells,
                "attributeDefinitions": len(definitions),
                "technicalDefinitions": sum(
                    1 for item in definitions if item["kind"] == "technical"
                ),
                "skippedMissingSku": skipped_missing_sku,
            },
        }
        return manifest
    finally:
        workbook.close()


def write_manifest(manifest: dict[str, object], output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(
        json.dumps(manifest, ensure_ascii=False, indent=2, sort_keys=False) + "\n",
        encoding="utf-8",
    )


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Build a deterministic Marco product-attribute repair manifest"
    )
    parser.add_argument("--xlsx", required=True, help="Path to Marco.xlsx")
    parser.add_argument("--out", required=True, help="Output JSON manifest path")
    parser.add_argument("--sheet", default="Worksheet", help="Worksheet name")
    args = parser.parse_args()

    xlsx_path = Path(args.xlsx)
    if not xlsx_path.is_file():
        raise FileNotFoundError(f"Excel source not found: {xlsx_path}")
    output_path = Path(args.out)
    manifest = build_manifest(xlsx_path, args.sheet)
    write_manifest(manifest, output_path)
    print(json.dumps(manifest["stats"], ensure_ascii=False, indent=2))
    print(f"Wrote {output_path}")


if __name__ == "__main__":
    main()
