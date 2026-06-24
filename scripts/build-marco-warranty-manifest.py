"""Build JSON manifest for backfill-product-warranty.cjs from Marco Excel."""

from __future__ import annotations

import argparse
import json
import re
from datetime import datetime, timezone
from pathlib import Path

import openpyxl


def norm_legacy_id(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def norm_sku(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def norm_cell(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def is_draft(value: object) -> bool:
    return str(value or "").strip().lower() in ("черновик", "draft")


def normalize_warranty_years(value: object) -> int | None:
    raw = norm_cell(value)
    if not raw:
        return None
    lowered = raw.lower()
    match = re.search(r"(\d+)", lowered)
    if not match:
        return None
    years = int(match.group(1))
    if years in (1, 2, 3):
        return years
    return None


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "--xlsx",
        default="Products for import/Marco - Redy for import.xlsx",
        help="Path to Marco Excel",
    )
    parser.add_argument(
        "--out",
        default="docs/reports/marco-warranty-manifest.json",
        help="Output manifest JSON",
    )
    parser.add_argument("--sheet", default="Worksheet", help="Sheet name")
    args = parser.parse_args()

    xlsx_path = Path(args.xlsx)
    out_path = Path(args.out)
    out_path.parent.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    if args.sheet not in wb.sheetnames:
        raise ValueError(f"Sheet '{args.sheet}' not found. Available: {', '.join(wb.sheetnames)}")

    ws = wb[args.sheet]
    rows = list(ws.iter_rows(min_row=1, values_only=True))
    headers = [str(h).strip() if h else "" for h in rows[0]]

    id_idx = headers.index("ID")
    name_idx = next(i for i, h in enumerate(headers) if h in ("Name", "ƒ"))
    sku_idx = next(i for i, h in enumerate(headers) if h in ("SKU", "Артикул", "Արտիկուլ"))
    draft_idx = headers.index("Черновик")
    warranty_idx = next(i for i, h in enumerate(headers) if h in ("Warranty", "Երաշխիք"))

    entries: list[dict[str, object]] = []
    warranty_counts = {1: 0, 2: 0, 3: 0, "none": 0}

    for row in rows[1:]:
        if not any(c is not None and str(c).strip() for c in row):
            continue
        legacy_id = norm_legacy_id(row[id_idx])
        if not legacy_id:
            continue
        warranty_years = normalize_warranty_years(row[warranty_idx])
        if warranty_years is None:
            warranty_counts["none"] += 1
        else:
            warranty_counts[warranty_years] += 1

        entries.append(
            {
                "legacyId": legacy_id,
                "sku": norm_sku(row[sku_idx]),
                "name": norm_cell(row[name_idx]),
                "isDraft": is_draft(row[draft_idx]),
                "warrantyYears": warranty_years,
            }
        )

    manifest = {
        "generatedAt": datetime.now(timezone.utc).isoformat(),
        "sourceXlsx": str(xlsx_path),
        "sheet": args.sheet,
        "entries": entries,
        "stats": {
            "rows": len(entries),
            "warrantyYears1": warranty_counts[1],
            "warrantyYears2": warranty_counts[2],
            "warrantyYears3": warranty_counts[3],
            "warrantyNone": warranty_counts["none"],
        },
    }

    out_path.write_text(json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(manifest["stats"], indent=2))
    print(f"Wrote {out_path}")


if __name__ == "__main__":
    main()
