import argparse
import csv
from pathlib import Path

import openpyxl


def normalize_header(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip().replace("\n", " ")


def normalize_cell(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def main() -> None:
    parser = argparse.ArgumentParser(description="Convert Marco XLSX into import-compatible CSV.")
    parser.add_argument("--xlsx", required=True, help="Absolute path to source XLSX")
    parser.add_argument("--csv", required=True, help="Absolute path to destination CSV")
    parser.add_argument("--sheet", default="Worksheet", help="Sheet name to read")
    args = parser.parse_args()

    xlsx_path = Path(args.xlsx)
    csv_path = Path(args.csv)

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    if args.sheet not in wb.sheetnames:
        raise ValueError(f"Sheet '{args.sheet}' not found. Available: {', '.join(wb.sheetnames)}")
    ws = wb[args.sheet]

    headers = [normalize_header(ws.cell(row=1, column=col).value) for col in range(1, ws.max_column + 1)]
    header_index = {header: idx for idx, header in enumerate(headers)}

    def pick(row_values: list[str], *keys: str) -> str:
        for key in keys:
            idx = header_index.get(key)
            if idx is None:
                continue
            value = row_values[idx]
            if value != "":
                return value
        return ""

    base_out_headers = [
        "ID",
        "Name",
        "SKU",
        "Short description",
        "Description",
        "price",
        "Sale price",
        "Category",
        "Brand",
        "Color",
        "Type",
        "Warranty",
        "Images",
    ]

    excluded_headers = {
        "",
        "ID",
        "Name",
        "SKU",
        "Артикул",
        "Արտիկուլ",
        "Տիպ Մեծածախ /  Մանրածախ",
        "Тип Мեծածախ /  Մանրածախ",
        "Type",
        "Черновик",
        "Description",
        "Desctriptop",
        "price",
        "Price",
        "Sale price",
        "sale price",
        "Category",
        "Brand",
        "Երաշխիք",
        "Warranty",
        "Color",
        "Images",
    }
    filter_source_headers = [h for h in headers if h not in excluded_headers]
    filter_headers = [f"Filter{idx + 1} - {header}" for idx, header in enumerate(filter_source_headers)]
    output_headers = base_out_headers + filter_headers

    csv_path.parent.mkdir(parents=True, exist_ok=True)
    with csv_path.open("w", newline="", encoding="utf-8") as fp:
        writer = csv.DictWriter(fp, fieldnames=output_headers, extrasaction="ignore")
        writer.writeheader()

        for row_idx in range(2, ws.max_row + 1):
            raw_row = [normalize_cell(ws.cell(row=row_idx, column=col).value) for col in range(1, ws.max_column + 1)]
            output_row = {
                "ID": pick(raw_row, "ID"),
                "Name": pick(raw_row, "Name"),
                "SKU": pick(raw_row, "SKU", "Артикул", "Արտիկուլ"),
                "Short description": pick(raw_row, "Short description", "Черновик"),
                "Description": pick(raw_row, "Description", "Desctriptop"),
                "price": pick(raw_row, "price", "Price"),
                "Sale price": pick(raw_row, "Sale price", "sale price"),
                "Category": pick(raw_row, "Category"),
                "Brand": pick(raw_row, "Brand"),
                "Color": pick(raw_row, "Color"),
                "Type": pick(raw_row, "Type", "Тип Մեծածախ /  Մանրածախ", "Տիպ Մեծածախ /  Մանրածախ"),
                "Warranty": pick(raw_row, "Warranty", "Երաշխիք"),
                "Images": pick(raw_row, "Images"),
            }

            for idx, source_header in enumerate(filter_source_headers):
                source_idx = header_index.get(source_header)
                output_row[f"Filter{idx + 1} - {source_header}"] = (
                    raw_row[source_idx] if source_idx is not None else ""
                )

            writer.writerow(output_row)

    print(f"Converted sheet '{args.sheet}' -> {csv_path}")
    print(f"Rows exported: {ws.max_row - 1}")
    print(f"Filter columns exported: {len(filter_headers)}")


if __name__ == "__main__":
    main()
