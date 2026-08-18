from __future__ import annotations

import importlib.util
import tempfile
import unittest
from pathlib import Path

import openpyxl
from openpyxl.styles import PatternFill


SCRIPT_PATH = Path(__file__).parents[1] / "build-marco-attribute-manifest.py"
SPEC = importlib.util.spec_from_file_location("build_marco_attribute_manifest", SCRIPT_PATH)
if SPEC is None or SPEC.loader is None:
    raise RuntimeError(f"Unable to import {SCRIPT_PATH}")
MODULE = importlib.util.module_from_spec(SPEC)
SPEC.loader.exec_module(MODULE)


class BuildMarcoAttributeManifestTests(unittest.TestCase):
    def make_workbook(self, path: Path) -> None:
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "Worksheet"
        worksheet.append(
            [
                "ID",
                "ƒ",
                "Артикул",
                "Price",
                "Color",
                "Օգտակար ծավալ",
                "Formatted blank spec",
            ]
        )

        worksheet.append([1, "Leading zero", 7, 100, "Սպիտակ", 42, None])
        worksheet["C2"].number_format = "00000"
        worksheet["G2"].number_format = "0.00"
        worksheet["G2"].fill = PatternFill(fill_type="solid", fgColor="FFFF00")

        worksheet.append([26105, "Simfer", "09468", 200, " Սև ", 60.0, ""])
        workbook.save(path)

    def test_extracts_09468_and_preserves_numeric_leading_zero_skus(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            xlsx_path = Path(temp_dir) / "Marco.xlsx"
            self.make_workbook(xlsx_path)

            manifest = MODULE.build_manifest(xlsx_path)

            self.assertEqual([entry["sku"] for entry in manifest["entries"]], ["00007", "09468"])
            by_sku = {entry["sku"]: entry for entry in manifest["entries"]}
            self.assertEqual(
                by_sku["09468"]["values"],
                [
                    {"definitionId": "color", "value": "Սև", "sourceCell": "E3"},
                    {"definitionId": "technical:1", "value": "60", "sourceCell": "F3"},
                ],
            )
            self.assertNotIn("Formatted blank spec", str(by_sku["00007"]["values"]))
            self.assertEqual(manifest["stats"]["attributeCells"], 4)
            self.assertEqual(manifest["stats"]["technicalDefinitions"], 2)

    def test_manifest_is_deterministic_and_ignores_cell_formatting_for_blank_values(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            xlsx_path = Path(temp_dir) / "Marco.xlsx"
            self.make_workbook(xlsx_path)

            first = MODULE.build_manifest(xlsx_path)
            second = MODULE.build_manifest(xlsx_path)

            self.assertEqual(first, second)
            self.assertNotIn("generatedAt", first)
            for entry in first["entries"]:
                self.assertFalse(any(item["definitionId"] == "technical:2" for item in entry["values"]))

    def test_legacy_marco_filter_index_matches_historical_converter_numbering(self) -> None:
        definitions = MODULE.build_attribute_definitions(
            ["ID", "Խոշոր Մանր", "Short description", "Stock", "Color", "Օգտակար ծավալ"]
        )
        useful_volume = next(item for item in definitions if item["kind"] == "technical")

        self.assertEqual(useful_volume["filterIndex"], 1)
        self.assertEqual(useful_volume["legacyFilterIndex"], 3)
        self.assertEqual(useful_volume["compatibilityKeys"], ["marco_filter_3"])
        self.assertFalse(any(item["label"] == "Խոշոր Մանր" for item in definitions))

    def test_splits_and_deduplicates_comma_separated_colors(self) -> None:
        self.assertEqual(
            MODULE.split_cell_values(" Սև, Սպիտակ, սև ", "color"),
            ["Սև", "Սպիտակ"],
        )
        self.assertEqual(MODULE.split_cell_values("60,5", "technical"), ["60,5"])


if __name__ == "__main__":
    unittest.main()
