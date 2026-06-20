"""Build JSON manifest for marco-import-reconcile.cjs from Marco Excel."""

from __future__ import annotations

import argparse
import csv
import json
import re
from datetime import datetime, timezone
from pathlib import Path

import openpyxl

JUNK_BRAND_NAMES = [
    "SUS",
    "Vikas",
    "DEO",
    "LARA",
    "Weston",
    "Simmer",
    "Ugur",
    "Uz8787",
]


def norm_name(value: object) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value).strip()).lower()


def norm_sku(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def norm_legacy_id(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def is_draft(value: object) -> bool:
    return str(value or "").strip().lower() in ("черновик", "draft")


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "--xlsx",
        default="Products for import/Marco - Redy for import.xlsx",
        help="Path to Marco Excel",
    )
    parser.add_argument(
        "--out",
        default="docs/reports/marco-import-reconcile-manifest.json",
        help="Output manifest JSON",
    )
    parser.add_argument(
        "--articule-csv",
        default="Products for import/articules-from-no-articule-sheet.csv",
        help="CSV with legacy ID -> new articul from no articule sheet",
    )
    args = parser.parse_args()

    xlsx_path = Path(args.xlsx)
    articule_csv_path = Path(args.articule_csv)
    out_path = Path(args.out)
    out_path.parent.mkdir(parents=True, exist_ok=True)

    enriched_updates: list[dict[str, str]] = []
    if articule_csv_path.exists():
        with articule_csv_path.open(encoding="utf-8-sig", newline="") as fp:
            reader = csv.DictReader(fp)
            for row in reader:
                legacy_id = norm_legacy_id(row.get("ID", ""))
                new_sku = norm_sku(row.get("New articul", ""))
                name = str(row.get("Name", "")).strip()
                if legacy_id and new_sku and name:
                    enriched_updates.append(
                        {"legacyId": legacy_id, "name": name, "newSku": new_sku}
                    )

    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)

    # no articule sheet kept for validation only
    ws_no = wb["no articule"]
    rows_no = list(ws_no.iter_rows(min_row=1, values_only=True))
    headers_no = [str(h).strip() if h else "" for h in rows_no[0]]
    name_idx_no = headers_no.index("Name")
    sku_idx_no = headers_no.index("Артикул")
    no_articule_count = 0
    for row in rows_no[1:]:
        if not any(c is not None and str(c).strip() for c in row):
            continue
        name = str(row[name_idx_no]).strip() if row[name_idx_no] is not None else ""
        new_sku = norm_sku(row[sku_idx_no])
        if name and new_sku:
            no_articule_count += 1

    ws = wb["Worksheet"]
    rows = list(ws.iter_rows(min_row=1, values_only=True))
    headers = [str(h).strip() if h else "" for h in rows[0]]
    id_idx = headers.index("ID")
    name_idx = headers.index("ƒ")
    sku_idx = headers.index("Артикул")
    brand_idx = headers.index("Brand")
    draft_idx = headers.index("Черновик")

    draft_skus: list[str] = []
    clear_brand_skus: list[str] = []

    for row in rows[1:]:
        if not any(c is not None and str(c).strip() for c in row):
            continue
        legacy_id = norm_legacy_id(row[id_idx])
        name = str(row[name_idx]).strip() if row[name_idx] is not None else ""
        sku = norm_sku(row[sku_idx])
        brand = str(row[brand_idx]).strip() if row[brand_idx] is not None else ""
        draft = row[draft_idx]

        if sku and is_draft(draft):
            draft_skus.append(sku)
        if sku and not brand:
            clear_brand_skus.append(sku)

    manifest = {
        "generatedAt": datetime.now(timezone.utc).isoformat(),
        "sourceXlsx": str(xlsx_path),
        "draftSkus": sorted(set(draft_skus)),
        "clearBrandSkus": sorted(set(clear_brand_skus)),
        "articuleUpdates": enriched_updates,
        "junkBrandNames": JUNK_BRAND_NAMES,
        "stats": {
            "draftSkus": len(set(draft_skus)),
            "clearBrandSkus": len(set(clear_brand_skus)),
            "articuleUpdates": len(enriched_updates),
            "noArticuleSheetRows": no_articule_count,
        },
    }

    out_path.write_text(json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(manifest["stats"], indent=2))
    print(f"Wrote {out_path}")


if __name__ == "__main__":
    main()
